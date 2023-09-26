 # Import libraries
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo

# Import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors

# Create window Tkinter
window = tk.Tk()
 
# Name our Tkinter application title
window.title(" BOM Explosion App ")
 
# Define window size in Tkinter python
window.geometry("600x450")

#Set filename as empty
filename = ""

#Read file
def select_file():
    global filename
    filetypes = (
        ('text files', '*.xlsx'),
        ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)

    showinfo(
        title='Selected File',
        message=filename
    )

open_button = ttk.Button(
    window,
    text='Open a File',
    command=select_file
)

open_button.pack(expand=True)

#Label of our program
label = tk.Label(window, text="Welcome to BOM Explosion Program",
font=('Calibri 15 bold'))
label.pack(pady=20)

#Define functionality of the button
def on_click_btn1():
    label["text"] = "Running"
    global filename
    if(filename == ""):
        label["text"] = "Please select a valid file"
        exit
    print(filename)
    wb = load_workbook(filename)
    ws1 = wb['Sheet1']
    wb.create_sheet("Explosion")

    ws1.delete_cols(2, 3)
    ws1.delete_cols(5, 11)
    wb.save(filename)

    wb = load_workbook(filename)
    ws1 = wb["Sheet1"]
    ws2 = wb["Sheet2"]
    we = wb['Explosion']

    print("Running...")

    count = 1
    spaces = []
    new_spaces = []
    no_of_s1_rows = ws1.max_row
    no_of_s2_rows = ws2.max_row
    fifty_done = False
    greenFill = PatternFill(start_color='80FF00', end_color='80FF00', fill_type='solid')
    itr_row = 1

    for row in range(1 , no_of_s2_rows):
        series = str(ws2["A" + str(row)].value)
        found = False
        row2_itr = 1
        for row2 in range(1 , no_of_s1_rows):
            if(series == str(ws1["B" + str(row2)].value)):
                found = True
                row2_itr = row2
        if(found == False):
            continue
        item_description = str(ws1["C" + str(row2_itr)].value)
        #print(type(series))
        #print(type("5"))
        level = ws1["A" + str(row2_itr)].value
        if((row2_itr + 1) > no_of_s1_rows):
            break
        level_next = ws1["A" + str(row2_itr+1)].value
        #print(level_next)
        next_count = row2_itr 
        level_count = row2_itr + 1
        exp_added = False
        while(level_next > level):
            next_count+=1
            if(level_next == (level + 1)):
                count+=1
                if(exp_added == False):
                    data_for = (series,item_description)
                    we.append(data_for)
                    itr_row+=1
                    count+=1
                    exp_added = True
                row_insert = ws1[next_count]
                row_with_values = [cell.value for cell in row_insert]
                itr_col = 3
                #print(itr_row)
                for item in row_with_values:
                    we.cell(row = itr_row, column = itr_col).value = item
                    itr_col+=1
                itr_row+=1
            level_count+=1
            level_next = ws1["A" + str(level_count)].value
            #print(level_next)
        spaces.append(count)
        #itr_row+=1
        if(count/no_of_s2_rows > 0.5 and fifty_done == False):
            print("Running...50% Done")
            #label["text"] = "Running...50% Done"
            fifty_done = True
        #print(count)
        count+=1
        new_spaces.append(itr_row)

    wb.save(filename)

    wb = load_workbook(filename)
    we = wb['Explosion']

    for x in new_spaces:
        we["A" + str(x)].fill=greenFill
        we["B" + str(x)].fill=greenFill

    for x in reversed(new_spaces):
        we.insert_rows(int(x))
        #print(x)


    we["A" + str(1)].fill=greenFill
    we["B" + str(1)].fill=greenFill

    wb.save(filename)

    print("100% Completed")
    label["text"] = "100% Completed"

btn1 = tk.Button(window, text="Perform Explosion", command=on_click_btn1)
btn1.pack(pady=20)

# Run main loop
window.mainloop()